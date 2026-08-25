"""Excel export (SS11.2).

The format the wedge actually uses. SS2.2: the larger population of manual QA
testers write test cases in Excel, Jira or Gherkin and execute them by hand, and
they are served badly or not at all -- so a `.feature` file is the proof of the
architecture and this is the thing a tester opens on Monday.

    One sheet per test case, plus a preconditions sheet, a parameters sheet
    and a warnings sheet.

The evidence column carries event ids rather than tool call ids. A spreadsheet
is where somebody *runs* the test; the audit trail for whether to believe it
lives in the trace sidecar and the review UI, and putting `tc_0447` in front of
a tester executing a step would be noise at exactly the wrong moment.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from server.config import ProjectConfig
from server.models import IRDocument, TestCaseIR
from server.pipeline.narrative import build_narrative
from server.renderers.base import ExportResult, review_warnings, test_cases

#: Excel rejects these in a sheet name, and silently truncates past 31 chars.
FORBIDDEN = re.compile(r"[\[\]:*?/\\]")
MAX_SHEET_NAME = 31

HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
MUTED = Font(italic=True, color="5B6470")

#: SS11.2 offers "one row per step in a flat sheet" as an alternative shape and
#: this takes it, with the two columns that decide what the file IS.
#:
#: Without `Pass / Fail` and `Notes` this is an export -- a record of what the
#: tool generated, which a tester reads once and closes. With them it is a test
#: script: something a person opens on Monday, works down, and fills in. The
#: larger half of SS2.2's wedge does not write Gherkin, and this is the file
#: they actually execute.
#:
#: They are left EMPTY. A generated pass mark would be a claim about a run
#: nobody made.
STEP_COLUMNS = [
    ("#", 5),
    ("Keyword", 10),
    ("Action", 52),
    ("Expected result", 40),
    ("Pass / Fail", 12),
    ("Notes", 30),
    ("Evidence", 18),
    ("Confidence", 12),
]

#: 1-based, and used in three places. Named so a column inserted before them
#: cannot silently write results into the wrong cells.
COL_EXPECTED = 4
COL_RESULT = 5
COL_EVIDENCE = 7
COL_CONFIDENCE = 8


class ExcelExporter:
    """One workbook per run, one sheet per test case."""

    name = "xlsx"

    def export(self, ir: IRDocument, *, out_dir: Path, config: ProjectConfig) -> ExportResult:
        workbook = Workbook()
        workbook.remove(workbook.active)

        used: set[str] = set()
        for case in test_cases(ir):
            sheet = workbook.create_sheet(_sheet_name(case, used))
            _write_case(sheet, case)

        _write_preconditions(workbook.create_sheet("Preconditions"), ir)
        _write_parameters(workbook.create_sheet("Parameters"), ir)
        _write_warnings(workbook.create_sheet("Warnings"), ir)
        _write_suggestions(workbook, ir)

        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / f"{ir.recordingId}.xlsx"
        workbook.save(path)

        return ExportResult(exporter=self.name, files=[path], warnings=review_warnings(ir))


# --------------------------------------------------------------------------
# sheets
# --------------------------------------------------------------------------


def _write_case(sheet: Worksheet, case: TestCaseIR) -> None:
    sheet["A1"] = case.scenarioName or case.title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = case.description or ""
    sheet["A2"].font = MUTED
    if case.objective:
        sheet["A3"] = f"Objective as stated by the tester: {case.objective}"
        sheet["A3"].font = MUTED

    _header(sheet, row=5, columns=STEP_COLUMNS)

    narrative = build_narrative(case.steps)
    row = 6
    number = 0

    for line in narrative.body:
        if line.is_assertion:
            # An expected result belongs beside the action that produced it,
            # not on a row of its own: a tester executing step 4 needs to see
            # what to check without scrolling to find it.
            cell = sheet.cell(row=row - 1, column=4)
            cell.value = f"{cell.value}\n{line.text}" if cell.value else line.text
            continue

        number += 1
        step = line.step
        sheet.cell(row=row, column=1, value=number)
        sheet.cell(row=row, column=2, value=line.keyword)
        sheet.cell(row=row, column=3, value=line.text)
        sheet.cell(row=row, column=COL_EXPECTED, value="")
        # Pass / Fail and Notes stay empty: they are for the person running
        # this, and anything written here by the tool would be a claim about a
        # run that has not happened.
        sheet.cell(row=row, column=COL_RESULT, value="")
        sheet.cell(row=row, column=COL_RESULT + 1, value="")
        sheet.cell(row=row, column=COL_EVIDENCE, value=_evidence(step))
        sheet.cell(row=row, column=COL_CONFIDENCE, value=step.confidence.value.title())

        if step.escalation:
            sheet.cell(
                row=row, column=COL_EXPECTED, value=f"[the agent asks] {step.escalation}"
            )
        row += 1

    _result_validation(sheet, first=6, last=max(row - 1, 6))

    for line in sheet.iter_rows(min_row=6, max_row=max(row - 1, 6), max_col=len(STEP_COLUMNS)):
        for cell in line:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if case.omitted:
        row += 1
        total = sum(o.eventCount for o in case.omitted)
        sheet.cell(
            row=row,
            column=1,
            value=(
                f"{total} action(s) were left out of this test case as exploratory or "
                f"abandoned. See the evidence sidecar."
            ),
        ).font = MUTED

    sheet.freeze_panes = "A6"


def _result_validation(sheet: Worksheet, *, first: int, last: int) -> None:
    """A dropdown on the Pass / Fail column.

    Not decoration. A results column somebody types into by hand comes back as
    "pass", "PASS", "ok", "y" and a tick character, and a suite whose results
    cannot be counted is a suite nobody reports on. Three values, chosen so
    "Blocked" exists -- a step that could not be run is not a failure, and
    forcing that choice is how a real run gets misreported.

    Best effort: an Excel that ignores the validation still shows a usable
    column, so a failure here must not cost the export.
    """
    if last < first:
        return
    try:
        from openpyxl.worksheet.datavalidation import DataValidation

        rule = DataValidation(
            type="list", formula1='"Pass,Fail,Blocked"', allow_blank=True, showDropDown=False
        )
        rule.error = "Choose Pass, Fail or Blocked."
        rule.prompt = "Leave blank until this step has actually been run."
        sheet.add_data_validation(rule)
        column = get_column_letter(COL_RESULT)
        rule.add(f"{column}{first}:{column}{last}")
    except Exception:  # noqa: BLE001 - a dropdown is never worth losing the file for
        return


def _write_preconditions(sheet: Worksheet, ir: IRDocument) -> None:
    _header(sheet, row=1, columns=[("Test case", 24), ("Precondition", 70), ("Shared", 10)])
    row = 2
    for case in test_cases(ir):
        for precondition in case.preconditions:
            sheet.cell(row=row, column=1, value=case.scenarioName or case.title)
            sheet.cell(row=row, column=2, value=precondition.text)
            sheet.cell(row=row, column=3, value="yes" if precondition.shared else "no")
            row += 1
    if row == 2:
        sheet.cell(row=2, column=1, value="No preconditions were identified.").font = MUTED


def _write_parameters(sheet: Worksheet, ir: IRDocument) -> None:
    sheet["A1"] = "Supply a real value for each of these before running the test."
    sheet["A1"].font = MUTED
    _header(
        sheet,
        row=3,
        columns=[("Placeholder", 24), ("Category", 18), ("Your value", 30), ("Notes", 40)],
    )

    row = 4
    seen: set[str] = set()
    for case in test_cases(ir):
        for parameter in case.parameters:
            if parameter.placeholder in seen:
                continue
            seen.add(parameter.placeholder)
            sheet.cell(row=row, column=1, value=parameter.placeholder)
            sheet.cell(row=row, column=2, value=parameter.category)
            sheet.cell(row=row, column=3, value="")
            sheet.cell(row=row, column=4, value=parameter.description or "")
            row += 1
    if row == 4:
        sheet.cell(row=4, column=1, value="This test needs no parameters.").font = MUTED


def _write_warnings(sheet: Worksheet, ir: IRDocument) -> None:
    # SS6.8 -- a tool that admits what it does not know stays trusted. In a
    # spreadsheet that means a sheet, not a colour.
    _header(sheet, row=1, columns=[("Severity", 12), ("Test case", 22), ("What", 84)])
    row = 2
    for case in test_cases(ir):
        for warning in case.warnings:
            sheet.cell(row=row, column=1, value=warning.severity.value)
            sheet.cell(row=row, column=2, value=case.scenarioName or case.title)
            sheet.cell(row=row, column=3, value=warning.message)
            row += 1
    for note in review_warnings(ir):
        sheet.cell(row=row, column=1, value="review")
        sheet.cell(row=row, column=3, value=note)
        row += 1
    if row == 2:
        sheet.cell(row=2, column=1, value="Nothing needs review.").font = MUTED


def _write_suggestions(workbook: Workbook, ir: IRDocument) -> None:
    """SS9.8 -- strictly quarantined.

    Its own sheet, labelled unverified, and never a row in a test case. A
    suggestion that reaches somebody as a step is the failure this separation
    exists to prevent.
    """
    suggestions = [(c, s) for c in ir.testCases for s in (c.suggestions or [])]
    if not suggestions:
        return

    sheet = workbook.create_sheet("Coverage (unverified)")
    sheet["A1"] = (
        "NOT part of any test case, and not grounded in a retrieval. These are things "
        "the recording revealed that nothing has exercised yet."
    )
    sheet["A1"].font = MUTED
    _header(sheet, row=3, columns=[("Category", 22), ("Suggestion", 60), ("Why", 60)])

    for row, (_case, suggestion) in enumerate(suggestions, start=4):
        sheet.cell(row=row, column=1, value=suggestion.category.value)
        sheet.cell(row=row, column=2, value=suggestion.text)
        sheet.cell(row=row, column=3, value=suggestion.rationale)


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------


def _header(sheet: Worksheet, *, row: int, columns: list[tuple[str, int]]) -> None:
    for index, (label, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        sheet.column_dimensions[get_column_letter(index)].width = width


def _evidence(step) -> str:
    ids = step.eventIds
    if len(ids) <= 2:
        return ", ".join(ids)
    return f"{ids[0]}-{ids[-1]}"


def _sheet_name(case: TestCaseIR, used: set[str]) -> str:
    """Excel truncates past 31 characters and rejects []:*?/\\ silently.

    Two test cases whose names collide after truncation would otherwise cost
    one of them its sheet, so uniqueness is enforced here rather than left to
    openpyxl.
    """
    base = FORBIDDEN.sub("", case.scenarioName or case.title or case.id).strip() or case.id
    name = base[:MAX_SHEET_NAME]

    suffix = 2
    while name.casefold() in used:
        tail = f" ({suffix})"
        name = base[: MAX_SHEET_NAME - len(tail)] + tail
        suffix += 1

    used.add(name.casefold())
    return name


__all__ = ["ExcelExporter"]
